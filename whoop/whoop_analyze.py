"""
WHOOP Analysis — Key Metrics + Rolling HRV-CV
Outputs: whoop-data/whoop_metrics.xlsx

Usage:
  python3 whoop/whoop_analyze.py                    # all available data
  python3 whoop/whoop_analyze.py --from 2025-01-01  # from a specific date
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import timezone

import pandas as pd

DATA_DIR    = Path(__file__).parent.parent / "whoop-data"
OUTPUT_FILE = DATA_DIR / "whoop_metrics.xlsx"


def load(name: str):
    path = DATA_DIR / f"{name}.json"
    if not path.exists():
        sys.exit(f"❌  {path} not found — run: python3 whoop/whoop_fetch.py")
    with open(path) as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(description="Generate WHOOP metrics Excel report")
    parser.add_argument("--from", dest="from_date", metavar="YYYY-MM-DD",
                        help="Only include data from this date onward (default: all)")
    args = parser.parse_args()

    recovery = load("recovery")
    cycles   = load("cycles")
    sleep    = load("sleep")

    # ── Build recovery dataframe ──────────────────────────────────────────────
    rec_rows = []
    for r in recovery:
        if r.get("score_state") != "SCORED":
            continue
        s = r["score"]
        rec_rows.append({
            "cycle_id":        r["cycle_id"],
            "date":            pd.to_datetime(r["created_at"], utc=True),
            "recovery_score":  s.get("recovery_score"),
            "hrv_rmssd_milli": s.get("hrv_rmssd_milli"),
            "resting_hr":      s.get("resting_heart_rate"),
            "spo2_pct":        s.get("spo2_percentage"),
            "skin_temp_c":     s.get("skin_temp_celsius"),
        })

    rec_df = pd.DataFrame(rec_rows)
    rec_df["date"] = rec_df["date"].dt.normalize().dt.tz_localize(None)
    rec_df = rec_df.sort_values("date").reset_index(drop=True)

    # ── Build cycle dataframe ─────────────────────────────────────────────────
    cyc_rows = []
    for c in cycles:
        if c.get("score_state") != "SCORED":
            continue
        s = c["score"]
        cyc_rows.append({
            "cycle_id":   c["id"],
            "strain":     s.get("strain"),
            "avg_hr":     s.get("average_heart_rate"),
            "max_hr":     s.get("max_heart_rate"),
            "kilojoules": s.get("kilojoule"),
            "calories":   round(s.get("kilojoule", 0) / 4.184),
        })

    cyc_df = pd.DataFrame(cyc_rows)

    # ── Build sleep dataframe (main sleep only, no naps) ──────────────────────
    slp_rows = []
    for s in sleep:
        if s.get("score_state") != "SCORED" or s.get("nap"):
            continue
        sc = s["score"]
        ss = sc.get("stage_summary", {})
        sn = sc.get("sleep_needed", {})

        total_sleep_ms = (
            ss.get("total_light_sleep_time_milli", 0)
            + ss.get("total_slow_wave_sleep_time_milli", 0)
            + ss.get("total_rem_sleep_time_milli", 0)
        )

        slp_rows.append({
            "cycle_id":               s["cycle_id"],
            "sleep_start":            s.get("start"),
            "sleep_end":              s.get("end"),
            "time_in_bed_hr":         round(ss.get("total_in_bed_time_milli", 0) / 3_600_000, 2),
            "total_sleep_hr":         round(total_sleep_ms / 3_600_000, 2),
            "light_sleep_hr":         round(ss.get("total_light_sleep_time_milli", 0) / 3_600_000, 2),
            "sws_hr":                 round(ss.get("total_slow_wave_sleep_time_milli", 0) / 3_600_000, 2),
            "rem_hr":                 round(ss.get("total_rem_sleep_time_milli", 0) / 3_600_000, 2),
            "awake_hr":               round(ss.get("total_awake_time_milli", 0) / 3_600_000, 2),
            "sleep_cycles":           ss.get("sleep_cycle_count"),
            "disturbances":           ss.get("disturbance_count"),
            "sleep_needed_hr":        round((sn.get("baseline_milli", 0) + sn.get("need_from_sleep_debt_milli", 0) + sn.get("need_from_recent_strain_milli", 0)) / 3_600_000, 2),
            "sleep_performance_pct":  sc.get("sleep_performance_percentage"),
            "sleep_consistency_pct":  sc.get("sleep_consistency_percentage"),
            "sleep_efficiency_pct":   sc.get("sleep_efficiency_percentage"),
            "respiratory_rate":       sc.get("respiratory_rate"),
        })

    slp_df = pd.DataFrame(slp_rows)

    # ── Merge all on cycle_id ─────────────────────────────────────────────────
    df = (
        rec_df
        .merge(cyc_df, on="cycle_id", how="left")
        .merge(slp_df, on="cycle_id", how="left")
    )

    if args.from_date:
        df = df[df["date"] >= args.from_date].copy()

    df = df.sort_values("date").reset_index(drop=True)

    if df.empty:
        sys.exit("❌  No data after applying date filter. Try a different --from date.")

    # ── Rolling HRV metrics ───────────────────────────────────────────────────
    df["hrv_7d_mean"]  = df["hrv_rmssd_milli"].rolling(7, min_periods=3).mean()
    df["hrv_7d_std"]   = df["hrv_rmssd_milli"].rolling(7, min_periods=3).std()
    df["hrv_cv_7d"]    = (df["hrv_7d_std"] / df["hrv_7d_mean"] * 100).round(2)
    df["hrv_30d_mean"] = df["hrv_rmssd_milli"].rolling(30, min_periods=7).mean()
    df["hrv_30d_std"]  = df["hrv_rmssd_milli"].rolling(30, min_periods=7).std()
    df["hrv_cv_30d"]   = (df["hrv_30d_std"] / df["hrv_30d_mean"] * 100).round(2)

    df["hrv_rmssd_milli"] = df["hrv_rmssd_milli"].round(2)
    df["hrv_7d_mean"]     = df["hrv_7d_mean"].round(2)
    df["hrv_30d_mean"]    = df["hrv_30d_mean"].round(2)

    # ── Select & rename columns ───────────────────────────────────────────────
    output_cols = {
        "date":                   "Date",
        "recovery_score":         "Recovery Score (%)",
        "hrv_rmssd_milli":        "HRV RMSSD (ms)",
        "hrv_7d_mean":            "HRV 7D Mean (ms)",
        "hrv_cv_7d":              "HRV-CV 7D (%)",
        "hrv_30d_mean":           "HRV 30D Mean (ms)",
        "hrv_cv_30d":             "HRV-CV 30D (%)",
        "resting_hr":             "Resting HR (bpm)",
        "spo2_pct":               "SpO2 (%)",
        "skin_temp_c":            "Skin Temp (°C)",
        "strain":                 "Day Strain",
        "avg_hr":                 "Avg HR (bpm)",
        "max_hr":                 "Max HR (bpm)",
        "calories":               "Calories",
        "total_sleep_hr":         "Total Sleep (hr)",
        "time_in_bed_hr":         "Time in Bed (hr)",
        "light_sleep_hr":         "Light Sleep (hr)",
        "sws_hr":                 "SWS (hr)",
        "rem_hr":                 "REM (hr)",
        "awake_hr":               "Awake in Bed (hr)",
        "sleep_cycles":           "Sleep Cycles",
        "disturbances":           "Disturbances",
        "sleep_needed_hr":        "Sleep Needed (hr)",
        "sleep_performance_pct":  "Sleep Performance (%)",
        "sleep_consistency_pct":  "Sleep Consistency (%)",
        "sleep_efficiency_pct":   "Sleep Efficiency (%)",
        "respiratory_rate":       "Respiratory Rate",
    }

    final_df = df[list(output_cols.keys())].rename(columns=output_cols)

    # ── Write Excel ───────────────────────────────────────────────────────────
    DATA_DIR.mkdir(exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Daily Metrics", index=False)

        wb = writer.book
        ws = writer.sheets["Daily Metrics"]

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule

        n_rows = len(final_df) + 1

        # Header styling
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="E0E0E0", size=10)

        for col_idx, col_name in enumerate(final_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="0F3460") if "HRV-CV" in col_name else header_fill

        col_widths = {
            "Date": 12, "Recovery Score (%)": 14, "HRV RMSSD (ms)": 14,
            "HRV 7D Mean (ms)": 15, "HRV-CV 7D (%)": 13,
            "HRV 30D Mean (ms)": 16, "HRV-CV 30D (%)": 14,
        }
        for col_idx, col_name in enumerate(final_df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 13)

        ws.row_dimensions[1].height = 32

        # Number formats
        col_map = {name: idx + 1 for idx, name in enumerate(final_df.columns)}
        pct_cols = ["Recovery Score (%)", "SpO2 (%)", "Sleep Performance (%)",
                    "Sleep Consistency (%)", "Sleep Efficiency (%)"]
        cv_cols  = ["HRV-CV 7D (%)", "HRV-CV 30D (%)"]

        for row in range(2, n_rows + 1):
            ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"

        for col_name, col_idx in col_map.items():
            fmt = None
            if col_name in pct_cols or col_name in cv_cols:
                fmt = "0.0"
            elif col_name in ("HRV RMSSD (ms)", "HRV 7D Mean (ms)", "HRV 30D Mean (ms)"):
                fmt = "0.00"
            elif col_name in ("Day Strain", "Respiratory Rate"):
                fmt = "0.0"
            elif col_name == "Calories":
                fmt = "#,##0"
            if fmt:
                for row in range(2, n_rows + 1):
                    ws.cell(row=row, column=col_idx).number_format = fmt

        # Color scales
        def data_range(c):
            return f"{get_column_letter(c)}2:{get_column_letter(c)}{n_rows}"

        ws.conditional_formatting.add(data_range(col_map["Recovery Score (%)"]), ColorScaleRule(
            start_type="num", start_value=0,  start_color="FF4C4C",
            mid_type="num",   mid_value=50,   mid_color="FFD700",
            end_type="num",   end_value=100,  end_color="00C851",
        ))
        for col in ["HRV RMSSD (ms)", "HRV 7D Mean (ms)", "HRV 30D Mean (ms)"]:
            ws.conditional_formatting.add(data_range(col_map[col]), ColorScaleRule(
                start_type="min", start_color="FF4C4C",
                mid_type="percentile", mid_value=50, mid_color="FFD700",
                end_type="max",  end_color="00C851",
            ))
        for col in ["HRV-CV 7D (%)", "HRV-CV 30D (%)"]:
            ws.conditional_formatting.add(data_range(col_map[col]), ColorScaleRule(
                start_type="min", start_color="00C851",
                mid_type="percentile", mid_value=50, mid_color="FFD700",
                end_type="max",  end_color="FF4C4C",
            ))
        for col, inv in [("Resting HR (bpm)", True), ("Day Strain", True), ("Sleep Performance (%)", False)]:
            rule = ColorScaleRule(
                start_type="min", start_color="00C851" if inv else "FF4C4C",
                mid_type="percentile", mid_value=50, mid_color="FFD700",
                end_type="max",  end_color="FF4C4C" if inv else "00C851",
            )
            ws.conditional_formatting.add(data_range(col_map[col]), rule)

        ws.freeze_panes = "B2"
        for row in ws.iter_rows(min_row=2, max_row=n_rows, min_col=1, max_col=len(final_df.columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Summary sheet
        sum_df = pd.DataFrame({
            "Metric": [
                "Days of Data", "Avg Recovery Score", "Avg HRV RMSSD (ms)",
                "Avg HRV-CV 7D (%)", "Avg HRV-CV 30D (%)", "Avg Resting HR (bpm)",
                "Avg SpO2 (%)", "Avg Day Strain", "Avg Total Sleep (hr)",
                "Avg Sleep Performance (%)", "Avg REM (hr)", "Avg SWS (hr)",
            ],
            "Value": [
                len(final_df),
                round(final_df["Recovery Score (%)"].mean(), 1),
                round(final_df["HRV RMSSD (ms)"].mean(), 2),
                round(final_df["HRV-CV 7D (%)"].mean(), 2),
                round(final_df["HRV-CV 30D (%)"].mean(), 2),
                round(final_df["Resting HR (bpm)"].mean(), 1),
                round(final_df["SpO2 (%)"].mean(), 2),
                round(final_df["Day Strain"].mean(), 1),
                round(final_df["Total Sleep (hr)"].mean(), 2),
                round(final_df["Sleep Performance (%)"].mean(), 1),
                round(final_df["REM (hr)"].mean(), 2),
                round(final_df["SWS (hr)"].mean(), 2),
            ]
        })
        sum_df.to_excel(writer, sheet_name="Summary", index=False)

        ws2 = writer.sheets["Summary"]
        for col_idx in [1, 2]:
            ws2.column_dimensions[get_column_letter(col_idx)].width = 28
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="E0E0E0")
            cell.fill = PatternFill("solid", fgColor="1A1A2E")
            cell.alignment = Alignment(horizontal="center")

    print(f"Saved → {OUTPUT_FILE}")
    print(f"Rows: {len(final_df)}  |  Columns: {len(final_df.columns)}")
    print()
    print("── Quick stats ──────────────────────────────────")
    print(f"  Date range        : {final_df['Date'].min().date()} → {final_df['Date'].max().date()}")
    print(f"  Avg Recovery      : {final_df['Recovery Score (%)'].mean():.1f}%")
    print(f"  Avg HRV RMSSD     : {final_df['HRV RMSSD (ms)'].mean():.1f} ms")
    print(f"  Avg HRV-CV 7D     : {final_df['HRV-CV 7D (%)'].mean():.1f}%")
    print(f"  Avg Resting HR    : {final_df['Resting HR (bpm)'].mean():.1f} bpm")
    print(f"  Avg Total Sleep   : {final_df['Total Sleep (hr)'].mean():.2f} hr")
    print(f"  Avg Day Strain    : {final_df['Day Strain'].mean():.1f}")


if __name__ == "__main__":
    main()
